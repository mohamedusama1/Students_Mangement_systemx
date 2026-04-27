"""
importpage.py — استيراد الطلاب والدرجات من CSV / Excel
========================================================
Tabs:
  1. Students Import  → first_student
  2. Civic Education  → civic_education_marks
  3. Arabic           → arabic_marks
"""

from customtkinter import *
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from time import strftime
import sqlite3, csv, os

try:
    import openpyxl
    EXCEL_SUPPORTED = True
except ImportError:
    EXCEL_SUPPORTED = False

import dashboard
from auth import session
from create_db import log_action

# ============================================================
#  SCHEMAS
# ============================================================

STUDENTS_COLS = [
    "student_name", "student_gender", "student_age",
    "address", "contact", "enrollment_date",
    "last_school", "repeated_years", "health_problem", "final_result"
]

GRADES_COLS = [
    "student_id", "student_name",
    "st_month", "nd_month", "chapter1", "half_year",
    "st_month1", "nd_month2", "chapter2", "chapter3",
    "final_exam", "final_result", "academic_year"
]

GENDER_MAP = {
    "ذكر":"Male","أنثى":"Female",
    "male":"Male","female":"Female",
    "Male":"Male","Female":"Female"
}

VALID_RESULTS = {"Excellent","Very Good","Good","Pass","Fail",
                 "ناجح","راسب","Pending"}


# ============================================================
#  MAIN CLASS
# ============================================================

class ImportClass:
    def __init__(self, root):
        self.root = root
        self.root.geometry('1200x720+80+20')
        self.root.title('Import Data')
        self.root.config(bg='white')
        self.root.resizable(False, False)

        if not session.require("add_student"):
            root.after(100, root.destroy)
            return

        self._build_ui()

    # ============================================================
    #  UI
    # ============================================================
    def _build_ui(self):
        root = self.root

        # ---- HEAD ----
        up = CTkFrame(root, width=1199, height=70, bg_color='#F6F5F5',
                      fg_color='#F6F5F5', border_color='#DA7297', border_width=2)
        up.place(x=1, y=1)
        Label(up, text='Import Data', font=('courier', 18, 'bold'),
              bg='#F6F5F5', fg='#DA7297').place(x=150, y=5, width=220, height=60)
        self._date_lbl = Label(up, font=('courier', 11, 'bold'), bg='#F6F5F5', fg='#DA7297')
        self._date_lbl.place(x=550, y=15, width=600, height=40)
        self._tick()

        def back():
            win = Toplevel()
            dashboard.DashboardClass(win)
            root.withdraw(); win.deiconify()

        CTkButton(up, text='←', width=100, height=68, fg_color='#DA7297',
                  text_color='white', bg_color='#F6F5F5', font=('arial', 30, 'bold'),
                  hover_color='#DA7297', corner_radius=0, command=back).place(x=2, y=2)

        # ---- MAIN NOTEBOOK (Students / Civic Ed / Arabic) ----
        main = CTkFrame(root, width=1197, height=648, bg_color='white',
                        fg_color='#F6F5F5', border_color='#DA7297', border_width=2)
        main.place(x=1, y=72)

        style = ttk.Style(); style.theme_use('clam')
        style.configure('TNotebook.Tab', font=('arial', 13, 'bold'),
                        padding=[16, 7], background='#DA7297', foreground='white')
        style.map('TNotebook.Tab', background=[('selected','#8B0000')])
        style.configure('Treeview', font=('arial', 11), rowheight=27, fieldbackground='white')
        style.configure('Treeview.Heading', background='#DA7297',
                        foreground='white', font=('arial', 11, 'bold'))

        nb = ttk.Notebook(main)
        nb.place(x=3, y=3, width=1190, height=640)

        # Students tab
        stu_tab = Frame(nb, bg='white')
        nb.add(stu_tab, text='👨‍🎓  Students')
        self.stu = _ImportTab(
            parent=stu_tab,
            schema=STUDENTS_COLS,
            table='first_student',
            validate_fn=self._validate_student,
            insert_fn=self._insert_student,
            template_row=["محمد أحمد","Male","10","القاهرة، مدينة نصر",
                          "01099887766","2024-09-01","مدرسة المستقبل","0","None","Pass"],
            col_widths={"student_name":140,"student_gender":70,"student_age":50,
                        "address":170,"contact":115,"enrollment_date":105,
                        "last_school":140,"repeated_years":75,
                        "health_problem":110,"final_result":80},
            col_labels={"student_name":"الاسم","student_gender":"الجنس",
                        "student_age":"السن","address":"العنوان",
                        "contact":"التليفون","enrollment_date":"تاريخ القيد",
                        "last_school":"المدرسة السابقة","repeated_years":"سنوات رسوب",
                        "health_problem":"مشكلة صحية","final_result":"النتيجة"},
        )

        # Civic Education tab
        civ_tab = Frame(nb, bg='white')
        nb.add(civ_tab, text='📚  Civic Education')
        self.civ = _ImportTab(
            parent=civ_tab,
            schema=GRADES_COLS,
            table='civic_education_marks',
            validate_fn=self._validate_grades,
            insert_fn=lambda rows, cur: self._insert_grades(rows, cur, 'civic_education_marks'),
            template_row=["1","بسمة سعيد","8","8","18","42",
                          "8","9","17","18","96","Excellent","2024-2025"],
            col_widths={"student_id":65,"student_name":130,
                        "st_month":65,"nd_month":65,"chapter1":70,"half_year":75,
                        "st_month1":65,"nd_month2":65,"chapter2":70,"chapter3":70,
                        "final_exam":80,"final_result":90,"academic_year":95},
            col_labels={"student_id":"ID طالب","student_name":"الاسم",
                        "st_month":"ش1","nd_month":"ش2","chapter1":"ف1",
                        "half_year":"نصف سنة","st_month1":"ش3","nd_month2":"ش4",
                        "chapter2":"ف2","chapter3":"ف3",
                        "final_exam":"نهائي","final_result":"النتيجة",
                        "academic_year":"العام الدراسي"},
        )

        # Arabic tab
        ara_tab = Frame(nb, bg='white')
        nb.add(ara_tab, text='📖  Arabic')
        self.ara = _ImportTab(
            parent=ara_tab,
            schema=GRADES_COLS,
            table='arabic_marks',
            validate_fn=self._validate_grades,
            insert_fn=lambda rows, cur: self._insert_grades(rows, cur, 'arabic_marks'),
            template_row=["1","بسمة سعيد","10","10","20","48",
                          "10","10","20","20","98","Excellent","2024-2025"],
            col_widths={"student_id":65,"student_name":130,
                        "st_month":65,"nd_month":65,"chapter1":70,"half_year":75,
                        "st_month1":65,"nd_month2":65,"chapter2":70,"chapter3":70,
                        "final_exam":80,"final_result":90,"academic_year":95},
            col_labels={"student_id":"ID طالب","student_name":"الاسم",
                        "st_month":"ش1","nd_month":"ش2","chapter1":"ف1",
                        "half_year":"نصف سنة","st_month1":"ش3","nd_month2":"ش4",
                        "chapter2":"ف2","chapter3":"ف3",
                        "final_exam":"نهائي","final_result":"النتيجة",
                        "academic_year":"العام الدراسي"},
        )

    def _tick(self):
        self._date_lbl.config(text=strftime('%I:%M:%S %p  -  %A  -  %d/%m/%Y'))
        self.root.after(1000, self._tick)

    # ============================================================
    #  VALIDATE — STUDENTS
    # ============================================================
    def _validate_student(self, row):
        errors, c = [], {}
        name = str(row.get('student_name', '')).strip()
        if not name: errors.append("الاسم فاضي")
        c['student_name'] = name

        gender = GENDER_MAP.get(str(row.get('student_gender', '')).strip())
        if not gender: errors.append(f"جنس غلط: {row.get('student_gender','')}")
        c['student_gender'] = gender or ''

        try:
            age = int(float(str(row.get('student_age', '')).strip()))
            if not 5 <= age <= 20: errors.append(f"سن خارج النطاق: {age}")
            c['student_age'] = age
        except ValueError:
            errors.append(f"سن مش رقم: {row.get('student_age','')}")
            c['student_age'] = 0

        c['address']         = str(row.get('address', '')).strip()
        c['enrollment_date'] = str(row.get('enrollment_date', '')).strip()
        c['last_school']     = str(row.get('last_school', '')).strip()

        contact = str(row.get('contact', '')).strip()
        if len(''.join(d for d in contact if d.isdigit())) < 10:
            errors.append(f"تليفون قصير: {contact}")
        c['contact'] = contact

        try:
            rep = int(float(str(row.get('repeated_years', '0')).strip()))
            if rep < 0: errors.append("سنوات رسوب سالبة")
            c['repeated_years'] = rep
        except ValueError:
            c['repeated_years'] = 0

        c['health_problem'] = str(row.get('health_problem', 'None')).strip() or 'None'
        c['final_result']   = str(row.get('final_result', 'Pending')).strip() or 'Pending'
        return errors, c

    # ============================================================
    #  VALIDATE — GRADES
    # ============================================================
    def _validate_grades(self, row):
        errors, c = [], {}

        try:
            sid = int(float(str(row.get('student_id', '')).strip()))
            if sid <= 0: errors.append("student_id لازم يكون موجب")
            c['student_id'] = sid
        except (ValueError, TypeError):
            errors.append(f"student_id مش رقم: {row.get('student_id','')}")
            c['student_id'] = 0

        c['student_name']  = str(row.get('student_name', '')).strip()
        c['academic_year'] = str(row.get('academic_year', '2024-2025')).strip() or '2024-2025'

        num_cols = ["st_month","nd_month","chapter1","half_year",
                    "st_month1","nd_month2","chapter2","chapter3","final_exam"]
        for col in num_cols:
            val = str(row.get(col, '0')).strip()
            try:
                fval = float(val)
                if fval < 0: errors.append(f"{col} سالب")
                c[col] = fval
            except ValueError:
                errors.append(f"{col} مش رقم: {val}")
                c[col] = 0.0

        result = str(row.get('final_result', '')).strip()
        c['final_result'] = result or 'Pending'
        return errors, c

    # ============================================================
    #  INSERT — STUDENTS
    # ============================================================
    def _insert_student(self, rows, cur):
        success = skipped = 0
        for c in rows:
            try:
                cur.execute("""
                    INSERT INTO first_student
                    (student_name,student_gender,student_age,address,contact,
                     enrollment_date,last_school,repeated_years,health_problem,final_result)
                    VALUES(?,?,?,?,?,?,?,?,?,?)
                """, (c['student_name'],c['student_gender'],c['student_age'],
                      c['address'],c['contact'],c['enrollment_date'],
                      c['last_school'],c['repeated_years'],
                      c['health_problem'],c['final_result']))
                success += 1
            except sqlite3.IntegrityError:
                skipped += 1
        return success, skipped

    # ============================================================
    #  INSERT — GRADES
    # ============================================================
    def _insert_grades(self, rows, cur, table):
        # تحقق إن student_ids موجودة في DB
        cur.execute("SELECT s_ID FROM first_student")
        valid_ids = {r[0] for r in cur.fetchall()}

        success = skipped = 0
        for c in rows:
            sid = c.get('student_id', 0)
            if sid not in valid_ids:
                skipped += 1
                continue
            try:
                cur.execute(f"""
                    INSERT INTO {table}
                    (student_id,student_name,st_month,nd_month,chapter1,half_year,
                     st_month1,nd_month2,chapter2,chapter3,final_exam,
                     final_result,academic_year)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (sid, c['student_name'],
                      c['st_month'],c['nd_month'],c['chapter1'],c['half_year'],
                      c['st_month1'],c['nd_month2'],c['chapter2'],c['chapter3'],
                      c['final_exam'],c['final_result'],c['academic_year']))
                success += 1
            except sqlite3.IntegrityError:
                skipped += 1
        return success, skipped


# ============================================================
#  IMPORT TAB — reusable widget per data type
# ============================================================

class _ImportTab:
    def __init__(self, parent, schema, table, validate_fn,
                 insert_fn, template_row, col_widths, col_labels):
        self.parent      = parent
        self.schema      = schema
        self.table       = table
        self.validate_fn = validate_fn
        self.insert_fn   = insert_fn
        self.template_row= template_row
        self.col_widths  = col_widths
        self.col_labels  = col_labels

        self.file_path  = None
        self.valid_rows = []
        self.error_rows = []

        self._build()

    def _build(self):
        p = self.parent

        # ---- Controls ----
        ctrl = CTkFrame(p, width=1185, height=120, fg_color='white',
                        bg_color='white', border_color='#DA7297', border_width=1)
        ctrl.pack(fill=X, padx=3, pady=3)

        self.file_var = StringVar(value="لم يتم اختيار ملف")
        CTkLabel(ctrl, text='الملف:', text_color='#DA7297',
                 font=('arial', 13, 'bold'), fg_color='white',
                 bg_color='white').place(x=10, y=12)
        CTkLabel(ctrl, textvariable=self.file_var, text_color='gray',
                 font=('arial', 11), fg_color='white', bg_color='white',
                 anchor='w', width=650).place(x=75, y=12)

        ext = "*.csv *.xlsx *.xls" if EXCEL_SUPPORTED else "*.csv"
        ftypes = [("CSV / Excel", ext), ("All files","*.*")]

        CTkButton(ctrl, text='📂  Browse', width=130, height=34,
                  fg_color='#DA7297', text_color='white', font=('arial', 12, 'bold'),
                  command=lambda: self._browse(ftypes)).place(x=10, y=50)

        CTkButton(ctrl, text='🔍  Preview', width=130, height=34,
                  fg_color='#DA7297', text_color='white', font=('arial', 12, 'bold'),
                  command=self._preview).place(x=150, y=50)

        self.import_btn = CTkButton(ctrl, text='✅  Import', width=150, height=34,
                                    fg_color='#2e7d32', text_color='white',
                                    font=('arial', 12, 'bold'),
                                    command=self._import, state='disabled')
        self.import_btn.place(x=290, y=50)

        CTkButton(ctrl, text='📥  Template', width=150, height=34,
                  fg_color='#666', text_color='white', font=('arial', 12, 'bold'),
                  command=self._download_template).place(x=450, y=50)

        # Stats
        self.s_total  = StringVar(value="إجمالي: 0")
        self.s_valid  = StringVar(value="صحيح: 0")
        self.s_errors = StringVar(value="أخطاء: 0")
        for i, (v, col) in enumerate([(self.s_total,'#333'),
                                       (self.s_valid,'#2e7d32'),
                                       (self.s_errors,'#c62828')]):
            Label(ctrl, textvariable=v, bg='white', fg=col,
                  font=('arial', 12, 'bold')).place(x=10 + i*180, y=95)

        # Progress
        self.prog_var = DoubleVar(value=0)
        self.prog_lbl = StringVar(value="")
        Label(ctrl, textvariable=self.prog_lbl, bg='white', fg='#DA7297',
              font=('arial', 10)).place(x=620, y=55)
        ttk.Progressbar(ctrl, variable=self.prog_var, maximum=100,
                        length=520, mode='determinate').place(x=620, y=80)

        # ---- Result Notebook ----
        nb = ttk.Notebook(p)
        nb.pack(fill=BOTH, expand=1, padx=3, pady=2)

        style = ttk.Style()
        style.configure('inner.TNotebook.Tab', font=('arial', 11),
                        padding=[10,4], background='#f8bbd0', foreground='#880e4f')

        valid_f = Frame(nb, bg='white'); nb.add(valid_f, text='✅  صحيح')
        err_f   = Frame(nb, bg='white'); nb.add(err_f,   text='❌  أخطاء')
        all_f   = Frame(nb, bg='white'); nb.add(all_f,   text='📋  الكل')

        self.v_tree = self._make_tree(valid_f, show_status=False)
        self.e_tree = self._make_tree(err_f,   show_status=True)
        self.a_tree = self._make_tree(all_f,   show_status=True)

    # ---- Tree builder ----
    def _make_tree(self, parent, show_status):
        cols = (["#","status"] if show_status else ["#"]) + self.schema
        f = Frame(parent, bg='gray')
        f.pack(fill=BOTH, expand=1, padx=2, pady=2)
        scx = Scrollbar(f, orient=HORIZONTAL)
        scy = Scrollbar(f, orient=VERTICAL)
        scx.pack(side=BOTTOM, fill=X)
        scy.pack(side=RIGHT, fill=Y)
        tree = ttk.Treeview(f, columns=cols, show='headings',
                            yscrollcommand=scy.set, xscrollcommand=scx.set)
        scy.config(command=tree.yview)
        scx.config(command=tree.xview)
        base_w = {"#":40,"status":200}
        for col in cols:
            w = base_w.get(col, self.col_widths.get(col, 90))
            lbl = {"#":"#","status":"الحالة"}.get(col, self.col_labels.get(col, col))
            tree.heading(col, text=lbl, anchor='center')
            tree.column(col, width=w, anchor='center')
        tree.tag_configure('valid', background='#e8f5e9')
        tree.tag_configure('error', background='#ffebee')
        tree.pack(fill=BOTH, expand=1)
        return tree

    # ---- Browse ----
    def _browse(self, ftypes):
        path = filedialog.askopenfilename(filetypes=ftypes)
        if not path: return
        self.file_path = path
        self.file_var.set(os.path.basename(path))
        self.valid_rows = []; self.error_rows = []
        for t in (self.v_tree, self.e_tree, self.a_tree):
            t.delete(*t.get_children())
        self.import_btn.configure(state='disabled')
        self.prog_var.set(0); self.prog_lbl.set("")

    # ---- Read file ----
    def _read_file(self, path):
        ext = os.path.splitext(path)[1].lower()
        rows = []
        if ext == '.csv':
            with open(path, encoding='utf-8-sig', newline='') as f:
                for row in csv.DictReader(f):
                    rows.append(dict(row))
        elif ext in ('.xlsx','.xls') and EXCEL_SUPPORTED:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            hdrs = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows())]
            for r in ws.iter_rows(min_row=2, values_only=True):
                rows.append({hdrs[j]: (str(v).strip() if v is not None else '')
                             for j, v in enumerate(r)})
            wb.close()
        else:
            raise ValueError(f"نوع الملف '{ext}' غير مدعوم.")
        return rows

    # ---- Preview ----
    def _preview(self):
        if not self.file_path:
            messagebox.showerror("خطأ", "اختار ملف أولاً."); return
        try:
            raw = self._read_file(self.file_path)
        except Exception as ex:
            messagebox.showerror("خطأ في القراءة", str(ex)); return
        if not raw:
            messagebox.showinfo("تنبيه", "الملف فاضي."); return

        self.valid_rows = []; self.error_rows = []
        all_entries = []
        for i, row in enumerate(raw, 1):
            errors, cleaned = self.validate_fn(row)
            entry = {"idx": i, "cleaned": cleaned, "errors": errors}
            all_entries.append(entry)
            (self.error_rows if errors else self.valid_rows).append(entry)

        self._fill_trees(all_entries)
        self.s_total.set(f"إجمالي: {len(raw)}")
        self.s_valid.set(f"صحيح: {len(self.valid_rows)}")
        self.s_errors.set(f"أخطاء: {len(self.error_rows)}")

        if self.valid_rows:
            self.import_btn.configure(state='normal')
        else:
            messagebox.showwarning("تنبيه", "مفيش صفوف صحيحة للاستيراد.")

    def _row_vals(self, entry, show_status):
        c = entry['cleaned']
        data = [c.get(col, '') for col in self.schema]
        if show_status:
            status = "✅ صحيح" if not entry['errors'] else "❌ " + " | ".join(entry['errors'])
            return [entry['idx'], status] + data
        return [entry['idx']] + data

    def _fill_trees(self, entries):
        for t in (self.v_tree, self.e_tree, self.a_tree):
            t.delete(*t.get_children())
        for e in self.valid_rows:
            self.v_tree.insert('', END, values=self._row_vals(e, False), tags=('valid',))
        for e in self.error_rows:
            self.e_tree.insert('', END, values=self._row_vals(e, True),  tags=('error',))
        for e in entries:
            tag = 'valid' if not e['errors'] else 'error'
            self.a_tree.insert('', END, values=self._row_vals(e, True), tags=(tag,))

    # ---- Import ----
    def _import(self):
        if not self.valid_rows:
            messagebox.showinfo("تنبيه", "مفيش صفوف صحيحة."); return
        if not messagebox.askyesno("تأكيد",
                f"هتستورد {len(self.valid_rows)} صف في جدول '{self.table}'.\n"
                f"الأخطاء ({len(self.error_rows)}) هيتم تجاهلها.\n\nمتأكد؟"):
            return

        self.import_btn.configure(state='disabled')
        total = len(self.valid_rows)
        cleaned_rows = [e['cleaned'] for e in self.valid_rows]

        try:
            con = sqlite3.connect('school.db')
            con.execute("PRAGMA foreign_keys = ON")
            cur = con.cursor()

            # insert in batches مع progress
            batch = 100
            success = skipped = 0
            for start in range(0, total, batch):
                chunk = cleaned_rows[start:start+batch]
                s, sk = self.insert_fn(chunk, cur)
                success += s; skipped += sk
                pct = min((start + batch) / total * 100, 100)
                self.prog_var.set(pct)
                self.prog_lbl.set(f"جارٍ الاستيراد... {min(start+batch,total)}/{total}")
                self.parent.update_idletasks()

            con.commit(); con.close()

            log_action(session.admin_username, f"IMPORT_{self.table.upper()}",
                       self.table, details=f"imported={success} skipped={skipped} "
                       f"file={os.path.basename(self.file_path)}")

            self.prog_var.set(100)
            self.prog_lbl.set(f"✅  تم استيراد {success} — تجاهل {skipped}")
            messagebox.showinfo("تم الاستيراد",
                f"✅  تم استيراد:  {success}  صف بنجاح\n"
                f"⏭  تم تجاهل:   {skipped}\n\n"
                f"الجدول:  {self.table}\n"
                f"الملف:   {os.path.basename(self.file_path)}")
        except Exception as ex:
            messagebox.showerror("خطأ", str(ex))
            self.import_btn.configure(state='normal')

    # ---- Template ----
    def _download_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension='.csv',
            filetypes=[("CSV","*.csv")],
            initialfile=f'{self.table}_template.csv')
        if not path: return
        try:
            with open(path, 'w', encoding='utf-8-sig', newline='') as f:
                w = csv.writer(f)
                w.writerow(self.schema)
                w.writerow(self.template_row)
            messagebox.showinfo("تم", f"تم حفظ القالب:\n{path}")
        except Exception as ex:
            messagebox.showerror("خطأ", str(ex))


if __name__ == "__main__":
    root = Tk()
    ImportClass(root)
    root.mainloop()
